import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def merge_and_align_cells(out_path, transformed_data):
    # Carrega o workbook usando openpyxl
    wb = load_workbook(out_path)
    ws = wb.active

    # Inicializa uma variável para controlar as linhas que serão mescladas
    start_row = 2  # assumindo que a primeira linha é o cabeçalho

    # Loop para percorrer as linhas e mesclar células nas colunas "COLUNA", "LINHA" e "TIPO_BLOCO"
    for i in range(3, len(transformed_data) + 2):
        # Verifica se o valor da célula atual é diferente do anterior para "COLUNA"
        if ws[f'B{i}'].value != ws[f'B{i-1}'].value:
            if start_row < i - 1:
                # Mescla as células na coluna "COLUNA"
                ws.merge_cells(f'B{start_row}:B{i-1}')
                
                # Mescla as células na coluna "LINHA" para o mesmo intervalo
                ws.merge_cells(f'C{start_row}:C{i-1}')
                
                # Mescla as células na coluna "TIPO_BLOCO" para o mesmo intervalo
                ws.merge_cells(f'H{start_row}:H{i-1}')
                
            start_row = i

    # Mescla as últimas células, se necessário
    if start_row < len(transformed_data) + 1:
        ws.merge_cells(f'B{start_row}:B{len(transformed_data)+1}')
        ws.merge_cells(f'C{start_row}:C{len(transformed_data)+1}')
        ws.merge_cells(f'H{start_row}:H{len(transformed_data)+1}')

    # Define o alinhamento vertical e horizontal para todas as células
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='center')

    # Salva o arquivo modificado
    wb.save(out_path)
    print('Arquivo salvo com sucesso')

def transform_pile_plant(data):
    # Formatar 'Position X', 'Position Y', e 'Position Z'
    data['Position X'] = data['Position X'].apply(lambda x: '{:.4f}'.format(x).replace('.', ','))
    data['Position Y'] = data['Position Y'].apply(lambda x: '{:.4f}'.format(x).replace('.', ','))
    data['Position Z'] = data['Position Z'].apply(lambda x: '{:.4f}'.format(x).replace('.', ','))
    
    # Criar a coluna 'ID_PILE'
    data['ID_PILE'] = data['C'].astype(str) + '-' + data['L'].astype(str) + data['P'].astype(str)
    
    # Reordenar as colunas para ter 'ID_PILE' primeiro
    cols = ['ID_PILE'] + ['C'] + ['L'] + ['P'] + ['ALTURA'] + ['Position X'] + ['Position Y'] + ['TIPO_BLOCO'] + ['Visibility1']
    data_id = data[cols]
    
    # Remover 'h=' da coluna 'ALTURA'
    data_id['ALTURA'] = data_id['ALTURA'].str.replace('h=', '') 
    
    # Manter apenas as colunas especificadas
    cols_to_keep = ['ID_PILE', 'Position X', 'Position Y', 'ALTURA', 'C', 'L', 'P', 'Visibility1','TIPO_BLOCO']
    cols_to_drop = [col for col in data_id.columns if col not in cols_to_keep]
    data_id_drop = data_id.drop(cols_to_drop, axis=1)
    
    # Renomear colunas
    data_renamed = data_id_drop.rename(columns={
        'Visibility1': 'TIPO_PILAR',
        'C': 'COLUNA', 
        'L': 'LINHA',
        'P': 'PILAR',
        'Position X': 'COORDENADA X',
        'Position Y': 'COORDENADA Y'
    }) 
    
    # Ordenar o DataFrame
    data_sorted = data_renamed.sort_values(by=['COLUNA', 'LINHA', 'PILAR'], ascending=[True, True, True])
    
    # Salvar o DataFrame resultante em um arquivo Excel
    out_path = "BLOCOS_PILE_PLANT_FINAL.xlsx"  # Caminho relativo
    data_sorted.to_excel(out_path, index=False)
    
    return out_path

# Interface do Streamlit
welcome = """
# DATA TRANSFORMATION

 Data Transformation for Pile Plant foundations files in Photovoltaics Power Plants
                            in .xlsx, .xls, or .csv format

"""

st.markdown(welcome)

upload_file = st.file_uploader("Insert your raw file:", type=["xlsx", "xls", "csv"])

if upload_file is not None:
    try:
        # Verificar a extensão do arquivo e ler o arquivo
        if upload_file.name.endswith('.xlsx'):
            data = pd.read_excel(upload_file, engine='openpyxl')
        elif upload_file.name.endswith('.xls'):
            data = pd.read_excel(upload_file, engine='xlrd')  # Usar 'xlrd' para arquivos .xls
        elif upload_file.name.endswith('.csv'):
            data = pd.read_csv(upload_file)

        # Exibir as primeiras linhas do DataFrame para verificação
        st.subheader("Data pre-visualization:")
        st.dataframe(data.head())

        if st.button("Transform Pile Plant Data"):
            # Chamar a função de transformação e obter o caminho do arquivo
            out_path = transform_pile_plant(data)

            # Chamar a função para mesclar e alinhar células
            merge_and_align_cells(out_path, pd.read_excel(out_path))

            # Ler o arquivo transformado para visualização
            transformed_data = pd.read_excel(out_path)

            st.subheader("Transformed Data Pre-Visualization")
            st.dataframe(transformed_data.head())

            # Oferecer o arquivo transformado para download
            with open(out_path, "rb") as file:
                st.download_button(
                    label="Download Transformed File",
                    data=file,
                    file_name="BLOCKS_PILE_PLANT_FINAL.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
else:
    st.info("Please upload a raw format file.")
