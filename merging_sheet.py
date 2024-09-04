import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Caminho do arquivo Excel
input_path = r"C:\\Users\\netlu\\OneDrive\\Área de Trabalho\\Nova pasta\\pile plant\\BLOCKS_PILE_PLANT_FINAL_v2.xlsx"
output_path = r"C:\\Users\\netlu\\OneDrive\\Área de Trabalho\\Nova pasta\\pile plant\\arquivo.xlsx"

# Carrega os dados do Excel
data_merge = pd.read_excel(input_path)

# Carrega o workbook usando openpyxl
wb = load_workbook(input_path)
ws = wb.active

# Inicializa uma variável para controlar as linhas que serão mescladas
start_row = 2  # assumindo que a primeira linha é o cabeçalho

# Loop para percorrer as linhas e mesclar células nas colunas "COLUNA", "LINHA" e "TIPO_BLOCO"
for i in range(3, len(data_merge) + 2):
    # Verifica se o valor da célula atual é diferente do anterior para "COLUNA"
    if ws[f'B{i}'].value != ws[f'B{i-1}'].value:
        if start_row < i - 1:
            # Mescla as células na coluna "COLUNA"
            ws.merge_cells(f'B{start_row}:B{i-1}')
            
            # Mescla as células na coluna "LINHA" para o mesmo intervalo
            ws.merge_cells(f'C{start_row}:C{i-1}')
            
            # Mescla as células na coluna "TIPO_BLOCO" para o mesmo intervalo
            ws.merge_cells(f'H{start_row}:H{i-1}')
            
        start_row = i

# Mescla as últimas células, se necessário
if start_row < len(data_merge) + 1:
    ws.merge_cells(f'B{start_row}:B{len(data_merge)+1}')
    ws.merge_cells(f'C{start_row}:C{len(data_merge)+1}')
    ws.merge_cells(f'H{start_row}:H{len(data_merge)+1}')

# Define o alinhamento vertical para todas as células
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical='center',horizontal='center')

# Salva o arquivo modificado
wb.save(output_path)
print('Arquivo salvo com sucesso')
